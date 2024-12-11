'''
Este archivo de python agrupa funciones para generar tabulados, gráficas y cualquier otra salida de los datos procesados y analizados.
'''

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Color, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
import os
import warnings
warnings.filterwarnings("ignore")

FUENTE_GLOBAL = 'Montserrat'

# función para establecer la fuente global
def establecer_fuente_global(fuente):
    global FUENTE_GLOBAL
    FUENTE_GLOBAL = fuente

# función para obtener la fuente global
def obtener_fuente_global():
    return FUENTE_GLOBAL

# --------------------------- # Clase DataFrameAnid # --------------------------- #
class DataFrameAnid(pd.DataFrame):
    """
    Clase personalizada que extiende `pd.DataFrame` para incluir metadatos adicionales.

    Esta clase permite crear un DataFrame con atributos personalizados que son útiles
    para almacenar información adicional sobre los datos, como título, subtítulo y notas
    explicativas que son relevantes para el conjunto de datos, pero no son parte de los
    datos mismos (tipo metadatos). 

    Parámetros
    ----------
    *args : list
        Argumentos variables posicionales pasados al constructor de `pd.DataFrame`.
    **kwargs : dict
        Argumentos variables nombrados pasados al constructor de `pd.DataFrame`.
        Además, se pueden especificar los siguientes argumentos para añadir metadatos:
        - titulo : str
            Título descriptivo del DataFrame.
        - subtitulo : str
            Subtítulo para proporcionar más contexto o información.
        - nota_a : str
            Notas adicionales (a-e) para información sobre los datos.
        - fuente : str
            La fuente de los datos.

    Atributos
    ----------
    _metadata : list of str
        Lista que contiene los nombres de los atributos de metadatos personalizados.
    titulo : str
        Título del DataFrame.
    subtitulo : str
        Subtítulo del DataFrame.
    nota_a : str
        Nota explicativa 'a'.
    nota_b : str
        Nota explicativa 'b', (por defecto) relacionada con el error estándar.
    nota_c : str
        Nota explicativa 'c', (por defecto) relacionada con el límite inferior del IC al 95%.
    nota_d : str
        Nota explicativa 'd', (por defecto) relacionada con el límite superior del IC al 95%.
    nota_e : str
        Nota explicativa 'e', (por defecto) relacionada con el coeficiente de variación.
    fuente : str
        La fuente de los datos del DataFrame.

    Métodos
    -------
    __finalize__(self, other, method=None, **kwargs)
        Método para propagar los metadatos personalizados cuando se realizan operaciones
        que devuelven un nuevo DataFrame.
    """
    _metadata = ['titulo', 'subtitulo', 'nota_a', 'nota_b', 'nota_c', 'nota_d', 'nota_e', 'fuente']

    def __init__(self, *args, **kwargs):
        self.titulo = kwargs.pop('titulo', 'CONASAMA. Encuesta Nacional de Salud Mental y Adicciones 2024 (ENASAMA).')
        self.subtitulo = kwargs.pop('subtitulo', '')
        self.nota_a = kwargs.pop('nota_a', '')
        self.nota_b = kwargs.pop('nota_b', '(2) Error estándar: calcula cuánto se aparta el valor obtenido  (estimación puntual) con respecto a la media (estimación puntual promedio) de la población.')
        self.nota_c = kwargs.pop('nota_c', '(3) Límite inferior del intervalo de confianza (IC inferior) al 95%.')
        self.nota_d = kwargs.pop('nota_d', '(4) Límite superior del intervalo de confianza (IC superior) al 95%.')
        self.nota_e = kwargs.pop('nota_e', '(5) Coeficiente de variación: es la proporción entre la desviación estándar y la media poblacional multiplicado por 100.')
        self.fuente = kwargs.pop('fuente', 'Fuente: CONASAMA, Conahcyt. Encuesta Nacional de Salud Mental y Adicciones, 2024.')

        super().__init__(*args, **kwargs)

    def __finalize__(self, other, method=None, **kwargs):
        """
        Propaga los metadatos personalizados a través de operaciones de pandas.

        Este método es invocado internamente por métodos de pandas para preservar metadatos en un nuevo DataFrame que es
        resultado de dichas operaciones. 

        Parámetros
        ----------
        other : DataFrameAnid or pandas DataFrame
            El objeto DataFrame a partir del cual se están propagando los metadatos.
        method : str, opcional
            El nombre del método que está invocando `__finalize__`. Esto puede ser útil si la lógica de propagación
            de metadatos necesita ser diferente dependiendo del método que se está ejecutando.
        **kwargs : dict
            Parámetros adicionales que pueden ser pasados a `__finalize__`. Estos pueden ser utilizados para ajustar 
            finamente la forma en que se copian los metadatos, aunque en la mayoría de los casos no son necesarios.

        Devoluciones
        -------
        self : DataFrameAnid
            Este método devuelve 'self' con los metadatos actualizados. Esto es importante ya que permite que el
            método sea utilizado en una cadena de métodos.

        Notas
        -----
        Este método es parte de la infraestructura interna de pandas para la gestión de metadatos.
        No es necesario invocar este método directamente; será llamado internamente por pandas cuando corresponda.

        Ejemplo
        -------
        Supongamos que `df1` y `df2` son instancias de `DataFrameAnid` con metadatos personalizados.
        Cuando se realiza una operación como `df1.append(df2)`, pandas llamará a `df1.__finalize__(df2)`
        para asegurarse de que los metadatos personalizados de `df1` o `df2` (dependiendo de la lógica de
        propagación) se transfieren al nuevo DataFrame resultante.
        """

        for name in self._metadata:
            object.__setattr__(self, name, getattr(other, name, ''))
        return self

# --------------------------- # Funciones para generar el reporte # --------------------------- #

def extraer_datos_y_atributos(df_anid):
    """
    Guarda los datos y atributos de un DataFrameAnid en un diccionario.

    Parámetros:
    -----------
    df_anid: DataFrameAnid
        DataFrameAnid con los datos y atributos a extraer.
    
    Regresa:
    --------
    dict
        Diccionario con los datos y atributos del DataFrameAnid.

    Nota:
    -----
    Los atributos deben coincidir con los establecidos en la clase DataFrameAnid.
    """
    datos_y_atributos = {
        'datos': df_anid,
        'titulo': df_anid.titulo,
        'subtitulo': df_anid.subtitulo,
        'nota_a': df_anid.nota_a,
        'nota_b': df_anid.nota_b,
        'nota_c': df_anid.nota_c,
        'nota_d': df_anid.nota_d,
        'nota_e': df_anid.nota_e,
        'fuente': df_anid.fuente       
    }
    return datos_y_atributos

def generar_diccionario_maestro(list_df):
    """ 
    Crea un diccionario que conjunta todos los Dataframes y sus atributos anidados en un solo diccionario.

    Parámetros:
    -----------
    list_df: list
        Lista que contiene cada uno de los DataFramesAnid que interese integrar en el archivo de salida.
    
    Regresa:
    --------
    dicc_maestro: dict
        Diccionario que contiene todos los DataFramesAnid y sus atributos anidados.
    """
    dicc_maestro = {}
    for i, df in enumerate(list_df):
        datos_y_atributos = extraer_datos_y_atributos(df)
        dicc_maestro[f"celula_{i}"] = datos_y_atributos
    
    return dicc_maestro

def generar_reporte(dicc, titulo_salida):
    """
    Genera un archivo de Excel con las tablas de los DataFramesAnid contenidos en un diccionario.

    Parámetros:
    -----------
    dicc: dict
        Diccionario que contiene todos los DataFramesAnid y sus atributos anidados.
    titulo_salida: str
        Nombre del archivo de salida.
    
    Regresa:
    --------
    None
        El archivo de Excel se guarda en la carpeta de datos procesados.
    """

    ruta_completa = os.path.join('..', 'datos', 'procesados','tab_consultas', titulo_salida + '.xlsx')

    with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:

        # crear la hoja de índice
        wb = writer.book
        indice_hoja = wb.create_sheet(title="Indice", index=0)

        # agrega los datos de cada df al archivo de Excel
        for num_tabla, (key, value) in enumerate(dicc.items(), start=1):
            value['datos'].to_excel(writer, sheet_name=f"Tabla {num_tabla}", startrow=3, index=True)

            ws = writer.sheets[f"Tabla {num_tabla}"]

            # añade el título y subtítulo
            ws.cell(row=1, column=1, value=value['titulo'])
            ws.cell(row=2, column=1, value=value['subtitulo'])

            # añade las notas al final del df
            u_fila = len(value['datos'].index) + 6
            ws.cell(row=u_fila, column=1, value=value['nota_a'])
            ws.cell(row=u_fila + 1, column=1, value=value['nota_b'])
            ws.cell(row=u_fila + 2, column=1, value=value['nota_c'])
            ws.cell(row=u_fila + 3, column=1, value=value['nota_d'])
            ws.cell(row=u_fila + 4, column=1, value=value['nota_e'])
            ws.cell(row=u_fila + 5, column=1, value=value['fuente'])

    return None

# --------------------------- # Clases de estilos tabulados # --------------------------- #
class Fuente:
    def __init__(self, name=None, size=None):
        self.name = name if name else obtener_fuente_global()
        self.size = size if size else 10

    def modificar_fuente(self, archivo):
        """ 
        Modifica la fuente de las celdas de un archivo excel.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel en el que se modificará la fuente.

        Regresa:
        ------------
        Archivo excel
        """
        
        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        fuente = Font(name=self.name, size=self.size)
        
        for hoja in wb:
            for row in hoja.iter_rows():
                for cell in row:
                    cell.font = fuente
    
        wb.save(os.path.join(ruta, archivo))

        return None
    
class Titulos:
    def __init__(self):
        pass

    def modificar_titulos(self, archivo):
        """ 
        Modifica la fuente de las celdas de un archivo excel.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel en el que se modificará la fuente.

        Regresa:
        ------------
        Archivo excel
        """
        
        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        fuente_global = obtener_fuente_global()
        fuente_titulo = Font(name=fuente_global, size=14, bold=True)
        fuente_subtitulo = Font(name=fuente_global, size=11, bold=False)

        for hoja in wb:
            if hoja.title == "Indice":
                continue

            for hoja in wb:
                for row in hoja.iter_rows():
                    for cell in row:
                        if cell.row == 1:
                            cell.font = fuente_titulo
                        elif cell.row == 2:
                            cell.font = fuente_subtitulo
    
        wb.save(os.path.join(ruta, archivo))

        return None
    
class Bordes:
    def __init__(self, no_border=Border(left=Side(style=None), 
                    right=Side(style=None),
                    top=Side(style=None),
                    bottom=Side(style=None))):
        self.no_border = no_border

    def modificar_bordes(self, archivo):
        """ 
        Modifica la fuente de las celdas de un archivo excel.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel en el que se modificará la fuente.

        Regresa:
        ------------
        Archivo excel
        """
        
        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        for hoja in wb:
            if hoja.title == "Indice":
                continue

            else:
                for hoja in wb:
                    for row in hoja.iter_rows():
                        for cell in row:
                            cell.border = self.no_border
    
        wb.save(os.path.join(ruta, archivo))

        return None

class NegritasHeaders:
    def __init__(self):
        pass

    def modificar_negritas(self, archivo):
        """ 
        Modifica la fuente de las celdas de un archivo excel.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel en el que se modificará la fuente.

        Regresa:
        ------------
        Archivo excel
        """
        
        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        fuente_global = obtener_fuente_global()
        negritas_headers = Font(name=fuente_global, size=10, bold=True)

        for hoja in wb:
            if hoja.title == "Indice":
                continue

            else:
                for hoja in wb:
                    for celda in hoja[4]:
                        celda.font = negritas_headers
    
        wb.save(os.path.join(ruta, archivo))

        return None

class NegritasIndex:
    def __init__(self):
        pass

    def iterar_negritas(self, archivo):
        """ 
        Itera por las filas de una hoja de excel para modificar las negritas de las celdas.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel que se tomará como referencia para modificar ciertas celdas en "negritas"
            (bold) para todas las pestañas, excepto la hoja de índice.
        
        Regresa:
        ------------
        Archivo excel
        """
        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        fuente_global = obtener_fuente_global()
        negritas_index = Font(name=fuente_global, size=10, bold=True)

        for hoja in wb:
            if hoja.title == "Indice":
                continue

            i_comienza = max(1, (hoja.max_row) - (hoja.max_row - 4))
            for i in range(i_comienza, hoja.max_row):
                if (i+1) % 3 == 0:
                    hoja[f'A{i}'].font = negritas_index

        wb.save(os.path.join(ruta, archivo))

        return None

class Notas:
    def __init__(self):
        pass

    def modificar_notas(self, archivo, num_notas=7):
        """ 
        Genera los cambios finales para las notas al final de cada hoja de excel.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel que se tomará como referencia para modificar las notas al final de cada hoja.
        num_notas: int, default=7
            Número de notas al final de cada hoja.

        Regresa:
        ------------
        Archivo excel
        """

        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        fuente_global = obtener_fuente_global()
        notas = Font(name=fuente_global, size=8, bold=False)

        for hoja in wb:
            if hoja.title == "Indice":
                continue
            else:
                for i in range(hoja.max_row, hoja.max_row-num_notas,-1):
                    hoja[f'A{i}'].font = notas
                
        wb.save(os.path.join(ruta, archivo))

        return None 

class AnchoCelda:
    def __init__(self, ancho_fijo_AB=20.57):
        self.ancho_fijo_AB = ancho_fijo_AB

    def fijar_ancho_celda(self, archivo):
        """ 
        Fija el ancho de las celdas de un archivo excel.

        Parámetros:
        ------------
        archivo: str
            nombre del archivo excel que se tomará como referencia para modificar el ancho de las celdas.

        Regresa:
        ------------
        Archivo excel
        """

        ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        wb = load_workbook(os.path.join(ruta, archivo))

        for hoja in wb:
            if hoja.title == "Indice":
                continue

            for col in hoja.iter_cols(min_row=4, max_row=4):
                column = col[0].column_letter

                if column == 'A' or column == 'B':
                    hoja.column_dimensions[column].width = self.ancho_fijo_AB
                else:
                    max_long = max(len(str(cell.value)) for cell in col if cell.value is not None)
                    ancho_ajustado = (max_long + 2) * 1.2
                    hoja.column_dimensions[column].width = ancho_ajustado

        wb.save(os.path.join(ruta, archivo))

        return None

class IndiceAgrupaciones:
    """
    Clase para generar el índice por agrupaciones de acuerdo el módulo al que pertenecen los resultados de cada tabla.

    Parámetros:
    ------------
    archivo: str
        nombre del archivo excel que se tomará como referencia para generar el índice.
    agrupaciones: dict
        diccionario que contiene las agrupaciones y las palabras clave que se utilizarán para identificar los módulos.
    
    Regresa:
    ------------
    Archivo excel
    """

    def __init__(self, archivo, agrupaciones):
        self.archivo = archivo
        self.agrupaciones = agrupaciones
        self.ruta = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
        self.archivo_completo = os.path.join(self.ruta, archivo)
        self.wb = load_workbook(self.archivo_completo)
        self.indice_hoja = self.wb["Indice"]

        fuente_global = obtener_fuente_global()
        self.fuente = Font(name=fuente_global, size=11, bold=False)
        self.fuente_titulo = Font(name=fuente_global, size=10, bold=True, color="1F497D")
        self.fuente_contenido = Font(name=fuente_global, size=11, bold=True)

    def aplicar_relleno_blanco(self):
        relleno_blanco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        max_fila = 500
        max_columna = 62

        for fila in range(1, max_fila + 1):
            for columna in range(1, max_columna + 1):
                celda = self.indice_hoja.cell(row=fila, column=columna)
                celda.fill = relleno_blanco

    def generar_indice(self):
        self.aplicar_relleno_blanco()

        for nombre_hoja in self.wb.sheetnames:
            if nombre_hoja != "Indice":
                titulo_hoja = self.wb[nombre_hoja]["A1"].value
                celda_titulo_indice = self.indice_hoja.cell(row=1, column=1, value=titulo_hoja)
                celda_titulo_indice.font = self.fuente_titulo
                break

        celda_contenido = self.indice_hoja.cell(row=3, column=2, value="Contenido")
        celda_contenido.font = self.fuente_contenido

        fila_actual = 5
        for grupo, palabras_clave in self.agrupaciones.items():
            filas_agregadas = 0

            for nombre_hoja in self.wb.sheetnames:
                if nombre_hoja == "Indice":
                    continue

                hoja = self.wb[nombre_hoja]
                subtítulo = hoja["A2"].value if hoja["A2"].value is not None else ""

                if any(palabra.lower() in subtítulo.lower() for palabra in palabras_clave):
                    if filas_agregadas == 0:
                        celda_grupo = self.indice_hoja.cell(row=fila_actual, column=3, value=grupo)
                        celda_grupo.font = self.fuente
                        fila_actual += 1

                    celda_subtítulo = self.indice_hoja.cell(row=fila_actual, column=4, value=subtítulo)
                    celda_subtítulo.hyperlink = Hyperlink(ref="", location=f"'{nombre_hoja}'!A1", tooltip=f"Ir a {nombre_hoja}")
                    celda_subtítulo.style = "Hyperlink"
                    celda_subtítulo.font = self.fuente
                    
                    fila_actual += 1
                    filas_agregadas += 1

            if filas_agregadas > 0:
                fila_actual += 1

        self.wb.save(self.archivo_completo)

# --------------------------- # Función para integrar todo el estilo al archivo xlsx  # --------------------------- #
def aplicar_estilo(nombre_archivo):
    """ 
    Aplica los estilos predetermindos a un archivo excel.

    Parámetros:
    ------------
    nombre_archivo: str
        nombre del archivo excel que se tomará como referencia para aplicar los estilos predeterminados.

    Regresa:
    ------------
    Archivo excel
    """

    ruta_completa = os.path.join('..', 'datos', 'procesados', 'tab_consultas')
    archivo = os.path.join(ruta_completa, nombre_archivo)

    # crear una instancia de la clase
    fuente_predeterminada = Fuente()
    titulos_predeterminados = Titulos()
    bordes_predeterminados = Bordes()
    negritas_headers = NegritasHeaders()
    negritas_index = NegritasIndex()
    notas = Notas()
    ancho_celda = AnchoCelda()

    # llamar al método para modificar el archivo
    fuente_predeterminada.modificar_fuente(nombre_archivo)
    titulos_predeterminados.modificar_titulos(nombre_archivo)
    bordes_predeterminados.modificar_bordes(nombre_archivo)
    negritas_headers.modificar_negritas(nombre_archivo)
    negritas_index.iterar_negritas(nombre_archivo)
    notas.modificar_notas(nombre_archivo)
    ancho_celda.fijar_ancho_celda(nombre_archivo)

    agrupaciones = {
    "Módulo Drogas": ["drogas", "ilícitas", 'médicas'],
    "Módulo Alcohol": ["alcohol"],
    "Módulo Tabaco": ["tabaco"],
    "Módulo Salud Mental": ["salud mental", "depresión", "ansiedad", "suicidio"],
    "Módulo de Actividad física": ["actividad física", "ejercicio", "deporte"]}

    indice = IndiceAgrupaciones(nombre_archivo, agrupaciones)
    indice.generar_indice()    

    return None

# --------------------------- # Función para dar formato de la salida csv a xlsx (Consulta) # --------------------------- #
def esquema_xlsx_consulta(d, cols_idx, cols_datos, i_sin_nombre=0, nombres_nuevos=None, redondeo=2):
    """
    Genera un DataFrame con la estructura de datos esperada para el reporte de consulta en formato xlsx.

    Parámetros
    ----------
    d : DataFrame
        DataFrame con los datos en formato de la salida csv.
    cols_idx : list de str
        Lista de nombres de las columnas que formarán el índice múltiple.
    cols_datos : list de str
        Lista de nombres de las columnas que contienen los datos.
    i_sin_nombre : int, opcional
        Número de columnas en el índice múltiple que no tendrán nombre. Por defecto es 0.
    nombres_nuevos : dict, opcional
        Diccionario adicional para renombrar los niveles del índice. Por defecto es None.
    redondeo : int, opcional
        Número de decimales para redondear los datos. Por defecto es 2.
    
    Regresa
    -------
    DataFrame
        DataFrame con la estructura de datos esperada para el reporte de consulta en formato xlsx.
    """
    
    # renombramientos fijo de las columnas con datos
    renombramientos_fijos = {
        'estimacion': 'Estimación puntual',
        'error_std': 'Error estándar',
        'ic_inf': 'IC Inferior',
        'ic_sup': 'IC Superior',
        'cv': 'Coeficiente de variación'
    }

    # índice múltiple
    i_sin_nombre = min(i_sin_nombre, len(cols_idx))
    nombres_indices = [None if i < i_sin_nombre else cols_idx[i] for i in range(len(cols_idx))]
    i = pd.MultiIndex.from_frame(d[cols_idx], names=nombres_indices)

    if nombres_nuevos:
        i.set_names([nombres_nuevos.get(name, name) for name in i.names], inplace=True)

    datos = d[cols_datos]
    df = pd.DataFrame(datos.values, index=i, columns=cols_datos)

    df = df.rename(columns=renombramientos_fijos)

    # redondear 
    df = df.round(redondeo)

    return df

# --------------------------- # Función para dar formato de la salida csv a csv/xlsx (Gema) # --------------------------- #

def esquema_csv_xlsx_gema(d):
    """
    Genera la estructura necesaria para generar capas en Gema.

    Parámetros:
    -----------
    d: DataFrame
        DataFrame con los datos a transformar.
    
    Regresa:
    --------
    DataFrame
        DataFrame con los datos transformados.
    """

    mapeo_sexo = {
        'Mujeres y hombres': 't',
        'Hombres': 'h',
        'Mujeres': 'm'
    }

    mapeo_grupo_etario = {
        'Población de 12-75 años': '12_75',
        '12-17 años': '12_17',
        '18-75 años': '18_75',
    }

    datos_transformados = []

    # para casos de desagregación "estatal"
    if 'clave_entidad_dai' in d.columns:
        g_id = 1

        for (cve, nom), grupo in d.groupby(['clave_entidad_dai', 'entidad']):
            fila = {'g_id': g_id, 'clave_entidad_dai': cve, 'entidad': nom}
            for _, row in grupo.iterrows():
                encabezado = f"{mapeo_sexo.get(row['sexo'], row['sexo'][0].lower())}_{mapeo_grupo_etario[row['grupo_etario']]}"
                fila[encabezado] = row['estimacion']
            datos_transformados.append(fila)
            g_id += 1

        datos_transformados_df = pd.DataFrame(datos_transformados).set_index(['g_id','clave_entidad_dai', 'entidad'])

    # para casos de desagregación "regional"

    elif 'nom_region' in d.columns:
        g_id = 1

        for reg, grupo in d.groupby('nom_region'):
            fila = {'g_id': g_id, 'nom_region': reg}
            for _, row in grupo.iterrows():
                encabezado = f"{mapeo_sexo.get(row['sexo'], row['sexo'][0].lower())}_{mapeo_grupo_etario[row['grupo_etario']]}"
                fila[encabezado] = row['estimacion']
            datos_transformados.append(fila)
            g_id += 1

        datos_transformados_df = pd.DataFrame(datos_transformados).set_index(['g_id','nom_region'])

    # para casos de desagregación por "estrato"

    elif 'estrato_tx' in d.columns:
        g_id = 1

        for sto, grupo in d.groupby('estrato_tx'):
            fila = {'g_id': g_id, 'estrato_tx': sto}
            for _, row in grupo.iterrows():
                encabezado = f"{mapeo_sexo.get(row['sexo'], row['sexo'][0].lower())}_{mapeo_grupo_etario[row['grupo_etario']]}"
                fila[encabezado] = row['estimacion']
            datos_transformados.append(fila)
            g_id += 1

        datos_transformados_df = pd.DataFrame(datos_transformados).set_index(['g_id','estrato_tx'])

    # para casos de desagregación "nacional"
    else:
        datos_transformados = {'g_id': 1}
        
        for idx, row in d.iterrows():
            encabezado = f"{mapeo_sexo.get(row['sexo'], row['sexo'][0].lower())}_{mapeo_grupo_etario[row['grupo_etario']]}"
            datos_transformados[encabezado] = row['estimacion']

        datos_transformados_df = pd.DataFrame([datos_transformados])

    return datos_transformados_df